import tkinter as tk
from tkinter import ttk
import re
import time
import csv
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import xlrd

version = 4.0
# window dimensions
HEIGHT = 700
WIDTH = 1000

'''
global variables
-userEntries is a list that will hold all of the user's inputs via the entry field
-counter is a counter variable used for functions that are recursively called
  (to check how many times we have called them)
-numCat is the total number of categories if the user selects manual entry
-catNames is a list of all of the category names to be applied to the KeywordCategories
-catThresh is a list of all of the category thresholds (number of total occurrences to pass that category)
-kw_dicts is a list that holds the KeywordCategory objects
-STAGE is a variable that stores what part of the program we are at. It is used as a "switch variable" (See getResponse documentation)
-textToBeAnalyzed is the students's text entry response.
-manual is a Boolean that is set based on whether the use wants to read from text files or manually input keywords & text
-csvInput is a list that stores the entries of each run of the text analysis so that it can be written to a csv file.
-PATH is the filepath for the files that are read and outputted (inside the lib folder)
-assignmentName is the name of hte Canvas assignment that we are grading.
-scores is a list of the marks given to the students on their text entries. See analyzetext() for marking
'''
userEntries = []
counter = 0
numCat = 0
catNames = []
catThresh = []
kw_dicts = []
STAGE = 0
textToBeAnalyzed = ''
manual = None
csvInput = []
PATH = f'Version {version}/lib/'
canvas_PATH = f'Version {version}/canvas/'
assignmentName = ''
scores = []
responseSpreadsheet = ''
df = None
readCSV = None
LINE_SPACE = '~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~'
welcome_message = f'Welcome to Keyword Finder Version {version}!' \
                  f'\n\nInstructions\n{LINE_SPACE}' \
                  f'\nTo use this program, you will either need the prepared spreadsheet (.xlsx file) with responses to the prompts, or to manually enter your responses to each prompt.' \
                  f'\nThe spreadsheet must be in the directory "{PATH}" for the program to work. If the file is not in this directory, please move it now.' \
                  f'\nOther important files that should be in that directory are: "keywords.txt" (the keyword file) & "key.ico" (the icon for this program).'\
                  f'\nYou should also have the directory "{canvas_PATH}". This directory should contain both a quiz export from Canvas and a gradebook export from canvas. Please note the name of these files in their respective cells in the spreadsheet. ' \
                  f'\nFor more information on formatting and directory organization, see the README.md file.\n{LINE_SPACE}\n'
'''
Tkinter Variables
The colour variables are used to colour the Tkinter GUI.
Here we are setting them to blank strings to be set later. 
-iconpath is the filepath concatenated with the icon file name (to be used to add the icon to the window)
'''
colourmode = ''
bgcolour = ''
textcolour = ''
scrlcolour = ''
framecolour = ''
iconpath = PATH + 'key.ico'


class KeywordCategory:
    '''
    KeywordCategory Class

    Variables:
    name- Category Name
    kwList - List/Array of Keywords
    keywords - Dictionary of Keywords with number of occurrences entry
    thresh - Integer threshold for how many keywords of this category are required for the text to "pass"

    Functions:
    Setters
    set_name, set_thresh - self explanatory
    set_list             - set kwList from a comma separated string
    set_list_fromFile    - set kwList from a given list
    set_dict             - set keywords Dictionary from a list/array

    Getters
    getName, getKeyList,
    getKeywords, getThresh  - self explanatory
    getPassFail             - counts number of occurrences stored in keywords Dictionary and
                               compares to threshold, if occurrences >= threshold, return True

    Other
    resetKeys               - Resets the occurrences for all keywords in Dictionary to 0 for next run
    print                   - defines how we want to print the KeywordCategories (to GUI and Console)
    textFileOutput          - defines formatting for outputting (and inputting) with text file.
    '''
    name = ''
    kwList = []
    keywords = {}
    thresh = 0

    def __int__(self):
        self.name = ''
        self.kwList = []
        self.keywords = {}
        self.thresh = 0

    def set_name(self, name):
        self.name = name

    def set_list(self, keywords):
        self.kwList = keywords.lower().split(',')

    def set_list_fromFile(self, keywordList):
        self.kwList = keywordList

    def set_dict(self, keywordList):
        self.keywords = dict.fromkeys(keywordList, 0)

    def set_thresh(self, thresh):
        self.thresh = thresh

    def resetKeys(self):
        keywords = self.getKeywords()
        for keys in keywords:
            keywords[keys] = 0
        self.set_list_fromFile(keywords)

    def getName(self):
        return self.name

    def getKeyList(self):
        return self.kwList

    def getKeywords(self):
        return self.keywords

    def getThresh(self):
        return self.thresh

    def getPassFail(self):
        count = 0
        kwDict = self.getKeywords()
        threshold = self.getThresh()
        for keys in kwDict:
            count += kwDict[keys]
        if count >= threshold:
            return True
        elif count < threshold:
            return False
        else:
            return None

    def print(self):
        return f'[\nName:{self.getName()}, \nthreshold:{str(self.getThresh())}, \nlist:{self.getKeyList()}, \ndict:{self.getKeywords()},\n]'

    def textFileOutput(self):
        keys = ''
        keylist = self.getKeyList()
        for i in range(0, len(keylist)):
            if i != len(keylist)-1:
                keys += keylist[i] + ','
            else:
                keys += keylist[i]
        retStr = '' + self.getName() + '\n' + str(self.getThresh()) + '\n' + keys + '\n'
        return retStr

# functions
def updateText(entry):
    '''
    This function updates the user's text upon the button being pressed or the Enter/ Return key being pressed
    The entry parameter is the text that is in the entry field at the time of button press/ enter key press
    If there is no text in the entry field, it is considered an invalid entry.
    If there is text, it is updated into the "convo" text field and the entry field is cleared
    The function then runs the getResponse() function which is what gets the next output of the computer.
    We pass the global variable STAGE into getResponse() so that it can determine which function to go to.
    '''
    global STAGE
    # if length > 0, there is text
    if len(entry) > 0:
        convo.config(state='normal')
        convo.insert(tk.END, '\nUser: ' + entry)
        userEntries.append(entry)
        user.delete(0, tk.END)
        print(f'userEntries list: {userEntries}') # Used to track entries that are stored & deleted
        getResponse(STAGE)
    else:
        convo.insert(tk.END, "\nComp: I'm sorry, that entry was invalid.")
    # GUI auto scroll down
    convo.see('end')
    convo.config(state='disable')


def getResponse(STAGE):
    '''
    This function takes in the global variable STAGE which this function
     uses to determine which part or "Stage" of the program we are on.
    - Initially we need to st up the keyword categories (stage = 0)
    - Then we create dictionaries from the keyword categories (stage = 1)
    - Then we get the input text that is going to be analyzed for keywords (stage = 2)
    - Then we find all of the keywords in the text. (stage = 3)
    '''
    if STAGE == 0:
        getKeywordCats()
    elif STAGE == 1:
        createDicts()
    elif STAGE == 2:
        getTextToAnalyze()
    elif STAGE == 3:
        analyzeText()


def getKeywordCats():
    '''
    This function gets all of the keywords and keyword categories from the user.
    The categories are stored as a custom Object type - KeywordCategories which can store the
        category name, keyword list (array), keyword dictionary, and threshold (see Documentation above)
    The global variables used are:
        userEntries, counter, numCat, catNames, catThresh kw_dicts, STAGE, manual, PATH, responseSpreadsheet, and df.
        - userEntries is the array that stores user input so we need this to check if the input was as expected
            and to account for yes or no (Y/N) responses.
        -counter is a counter variable, in this method it is used to count which category we are currently on.
        -numCat is the number of categories the user has requested.
        -catNames is the list storing the names of each category (to later set name var of KeywordCategories)
        -catThresh is the list storing the thresholds of each of the categories
        -kw_dicts is the list of all of the KeywordCategories.
        -STAGE is only used so that we can increment it at the end of this function
        -manual is a boolean that is set based on if user wants to read from text file or manual entry
        -PATH is the filepath where the files that are read/written to are stored
        -responseSpreadsheet is the filename for the .xlsx file that has predetermined responses to the prompts
        -df is the dataframe storing the responseSpreadsheet entries
    '''
    global userEntries, counter, numCat, catNames, catThresh, kw_dicts, STAGE, manual, PATH, responseSpreadsheet, df
    backUp = []
    count = len(userEntries)
    # count=0 means start, ask which method they want to use.
    if count == 0:
        convo.insert(tk.END, '\nComp: Would you like to use responses from spreadsheet or manual entry?'
                             '\nComp: (Enter "0" for spreadsheet or "1" manual entry)')

    # manual boolean decides which path to take
    elif count == 1 and manual is None:
        if userEntries[0] == '0':
            manual = False
            counter = 1
        elif userEntries[0] == '1':
            manual = True
        else:
            convo.insert(tk.END, f'\n\nComp: That entry was invalid. Please try again.')
            revert(1)

    # This portion is all manual user entry
    if manual:
        if count == 1:
            convo.insert(tk.END, '\nComp: How many keyword categories would you like?')
        # count=2 means they entered a number of categories, confirm the entry
        elif count == 2:
            # if the user entered a valid number of categories, it will be an integer
            try:
                numCat = int(userEntries[-1])
                if numCat == 1:
                    convo.insert(tk.END, f'\nComp: You have requested {numCat} category, is this correct? (Enter "Y" to confirm or "N" to change)')
                else:
                    convo.insert(tk.END, f'\nComp: You have requested {numCat} categories, is this correct? (Enter "Y" to confirm or "N" to change)')
            # if not, output error text and back up previous entries to go back
            except ValueError:
                convo.insert(tk.END, "\nComp: The number of categories must be an integer.")
                for i in range(0, len(userEntries)-1):
                    backUp.append(userEntries[i])
                userEntries.clear()
                userEntries = backUp.copy()
                count -= 1
                getResponse(STAGE)
        # count=3 means that we are now naming and setting threshold of the first category
        # unless 'n' was entered at count=1, then we reset the list and start again.
        elif count == 3:
            if userEntries[-1].lower() == 'y':  # If the user confirmed # categories, move on
                print(f'count = 3, counter == {counter}')
                convo.insert(tk.END, f'\nComp: Enter the name and threshold of category {(counter+1)} (In the form "name,threshold"):')
                counter += 1

            elif userEntries[(count-1)].lower() == 'n':
                userEntries.remove(userEntries[-2])
                userEntries.remove(userEntries[-1])
                counter = 0
                getResponse(STAGE)
            else:
                convo.insert(tk.END, "\nComp: I'm sorry, that entry was invalid. Please try again.")
                userEntries.remove(userEntries[-1])
                getResponse(STAGE)
        # This is the iterative step of asking for a category name and then confirming it,
        # So we need a generic conditional statement
        # numCat * 2 because there is a name/threshold and confirmation 'y' for each cat
        # 3 < and + 3 because this step starts after the first 3
        # We are selecting the name of a category
        # Unless 'n' was entered at the previous step, then we save the list up to the
        #  last category name and restart this step
        elif 3 < count < (numCat*2)+3 and counter < numCat and count % 2 != 0:
            if userEntries[-1].lower() == 'y':  # If the user confirmed # of categories or category name, move on
                convo.insert(tk.END, f'\nComp: Enter the category name and threshold of category {(counter+1)} (In the form "name,threshold"):')
                counter += 1
            elif userEntries[-1].lower() == 'n':
                for i in range(0, (count-2)):
                    backUp.append(userEntries[i])
                userEntries.clear()
                userEntries = backUp.copy()
                catNames.pop()
                catThresh.pop()
                counter -= 1
                getResponse(STAGE)
            else:
                convo.insert(tk.END, "\nComp: I'm sorry, that entry was invalid. Please try again.")
                revert(1)
                catNames.pop()
                catThresh.pop()
                count -= 1
                getResponse(STAGE)

        # Confirmation part of above step. Only accepts 'Y' or 'N' (lower or upper case)
        # Anything else will output "Invalid entry, try again"
        elif 3 < count < (numCat*2)+3 and counter <= numCat and count % 2 == 0:
            words = userEntries[-1].split(',')
            if len(words) > 1:
                # if user enters a valid name and threshold, the threshold should be an integer
                try:
                    convo.insert(tk.END, f'\nComp: Category {counter} is named "{words[0]}", '
                                            f' with a threshold of "{int(words[1])}"'
                                            f'\nComp: Is this correct? (Enter "Y" to confirm or "N" to change)')
                    catNames.append(words[0])
                    catThresh.append(int(words[1]))
                # if threshold is not an integer, ValueError is thrown
                except ValueError:
                    convo.insert(tk.END, "\nComp: The threshold must be an integer.")
                    revert(1)
                    counter -= 2
                    getResponse(STAGE)
            else:
                convo.insert(tk.END, "\nComp: Invalid Category name or threshold. Please try again.")
                revert(1)
                counter -= 2
                getResponse(STAGE)

        # When count = (numCat*2) + 3 that means that the user has inputted - see calc at top of 'iterative step'
        #  all categories and confirmed, lets confirm all categories to make sure.
        elif count == (numCat*2)+3 and counter >= numCat and userEntries[-1].lower() == 'y':
            c = 1  # temp variable
            for i in range(3, len(userEntries)-1, 2):
                words = userEntries[i].split(',')
                convo.insert(tk.END, f'\nComp: Category {c} is named "{words[0]}",'
                                        f'with a threshold of "{int(words[1])}"')
                c += 1
            convo.insert(tk.END, f'\nComp: Are these entries all correct? (Enter "Y" to confirm or "N" to change):')

        # Last entry is incorrect, backup list and go back
        elif count == (numCat*2) + 3 and counter >= numCat and userEntries[-1].lower() == 'n':
            for i in range(0, len(userEntries)-2):
                backUp.append(userEntries[i])
            userEntries.clear()
            userEntries = backUp.copy()
            catNames.pop()
            catThresh.pop()
            counter = numCat-1
            getResponse(STAGE)
        # At this point we are adding the KeywordCatgories to the list
        elif count > (numCat * 2) + 3:
            if userEntries[-1].lower() == 'y':
                for i in range(0, len(catNames)):
                    kw_dicts.append(KeywordCategory())
                # Set counter manual to None so that they can be reused
                counter = 0
                STAGE = 1
                getResponse(STAGE)
            elif userEntries[-1].lower() == 'n':
                backUp.append(userEntries[0])
                backUp.append(userEntries[1])
                userEntries.clear()
                userEntries = backUp.copy()
                counter = 0
                getResponse(STAGE)
            else:
                convo.insert(tk.END, "\nComp: I'm sorry, that entry was invalid. Please try again.")
                for i in range(0, len(userEntries) - 1):
                    backUp.append(userEntries[i])
                userEntries.clear()
                userEntries = backUp.copy()
                catNames.pop()
                catThresh.pop()
                count -= 1
                getResponse(STAGE)
        else:
            convo.insert(tk.END, "\nComp: I'm sorry, that entry was invalid.")
            for i in range(0, len(userEntries) - 1):
                backUp.append(userEntries[i])
            userEntries.clear()
            userEntries = backUp.copy()
            catNames.pop()
            catThresh.pop()
            count -= 1
            getResponse(STAGE)

    # If user wants to read responses from the spreadsheet
    if not manual and manual is not None:
        if counter == 1:
            convo.insert(tk.END, f'\nComp: Enter the filename of the spreadsheet (.xlsx file) you wish to use.')
            counter += 1

        elif counter == 2:
            responseSpreadsheet = userEntries[-1]
            if ".xlsx" not in responseSpreadsheet:
                responseSpreadsheet += '.xlsx'
            full_path = PATH + '' + responseSpreadsheet
            # preprocess method checks if filenames in .xlsx are valid and files are in the right place
            preprocess(full_path)
            try:
                # read .xlsx into dataframe
                df = pd.read_excel(full_path, index_col=None, na_values=['NA'], usecols='B')

                filename = df['Response'][0]
                fullname = PATH + '' + filename

                try:
                    file = open(fullname, 'r')
                    #convo.insert(tk.END, f'\nComp: From File:\n')
                    catNames = []  # stores all category names
                    catThresh = []  # stores all the category thresholds
                    kwlists = []  # stores all keywords

                    # For each line in the file, first line is category name
                    #  second line is threshold, third line is keyword list
                    for line in file:
                        line = line.replace('\n', '')
                        catNames.append(line)
                        catThresh.append(file.readline().replace('\n', ''))
                        kwlists.append(file.readline().replace('\n', ''))

                    try:
                        for i in range(0, len(catNames)):
                            kw_dicts.append(KeywordCategory())
                            kw_dicts[i].set_name(catNames[i])
                            kw_dicts[i].set_thresh(int(catThresh[i]))
                            kw_dicts[i].set_list(kwlists[i])
                            kw_dicts[i].set_dict(kw_dicts[i].getKeyList())
                            # convo.insert(tk.END, f'\nCategory{i + 1}:{kw_dicts[i].print()}')

                    # If threshold is not an integer, ValueError thrown
                    except ValueError:
                        convo.insert(tk.END, f'\nComp: There was a problem with the format of the text file.'
                                             f'\nComp: Please reformat it according to the README file, and try again.\n')

                    counter += 1
                    getResponse(STAGE)

                except FileNotFoundError:
                    convo.insert(tk.END, f'\nComp: Keyword file "{fullname}" not found in "{PATH}". '
                                         f'Please move the file to the directory or confirm you have entered the correct filename into the spreadsheet.\n')
                    counter -= 1
                    getResponse(STAGE)

            # If the spreadsheet name is incorrect or it is not in the right location, FNF Error. Jump to start.
            except FileNotFoundError:
                #print('Spreadsheet file not found!')
                getResponse(STAGE)

        # Move on to text to analyze
        elif counter == 3:
            manual = None
            counter = 0
            STAGE = 2
            getResponse(STAGE)

    convo.see('end')  # makes the scrollbar keep most recent messages on screen


def preprocess(full_path):
    '''
    This function preprocesses the entries in the prepared spreadsheet and checks if the files are in the correct directories.
    This is done using a Pandas dataframe ("df"). The responses are all put into a list for easy access.
    The file checking is done in a nested manner because each file is used in conjunction with the previous file,
      so if one file is gone, the next will not matter.
    The global variables used are: PATH, canvas_PATH, and df.
    -PATH and canvas_PATH are used to concatenate with the filenames to locate the files.
    -df is used to store the dataframe to be used for other functions to acess the reponses.
    '''
    global PATH, canvas_PATH, df
    responses = []

    try:
        df = pd.read_excel(full_path, index_col=None, na_values=['NA'], usecols='B')

    # If .xlsx does not exist or incorrect name was entered
    except FileNotFoundError:
        convo.insert(tk.END, f'\nComp: The file "{responseSpreadsheet}" could not be found in the directory "{PATH}". Please move the file to the directory or confirm you have entered the correct filename.\n')
        reset()

    else:
        # Add responses from file to list object
        for i in range(0, len(df['Response'])):
            responses.append(df['Response'][i])

        # Check for keywords file
        keywordsFile = responses[0]
        keyword_path = PATH + '' + keywordsFile

        try:
            open(keyword_path, 'r')
        except FileNotFoundError:
            convo.insert(tk.END, f'\nComp: The file "{keywordsFile}" could not be found in the directory "{keyword_path}". Please move the file to the directory or confirm you have entered the correct filename if you would like to read keywords from a spreadsheet.\n')
            reset()

        else:
            # Check for Canvas Quiz Export
            quizExport = responses[1]
            quizExport_path = canvas_PATH + '' + quizExport

            try:
                open(quizExport_path, 'r')
            except FileNotFoundError:
                convo.insert(tk.END, f'\nComp: The file "{quizExport}" could not be found in the directory "{canvas_PATH}". Please move the file to the directory or confirm you have entered the correct filename if you would like to read from the Canvas Quiz Export.\n')
                reset()

            else:
                # Check for Canvas Gradebook Export
                if responses[4].lower() == 'yes':
                    gradesExport = responses[5]
                    gradesExport_path = canvas_PATH + '' + gradesExport

                    try:
                        open(gradesExport_path, 'r')
                    except FileNotFoundError:
                        convo.insert(tk.END,
                                     f'\nComp: The file "{gradesExport}" could not be found in the directory "{canvas_PATH}". Please move the file to the directory or confirm you have entered the correct filename if you would like to export to the gradebook.\n')
                        reset()



def createDicts():
    '''
    This function creates the dictionaries of the keywords categories created in the getKeywordCats function.
    This is only used in the event of manual entry.
    This function uses the global variables: counter, kw_dicts, and STAGE
        -counter is used to count the category we are on, similar to in getKeywordCats
        -kw_dicts is the list (array) storing the KeywordCategory objects.
        -STAGE is updated when function is done so we can change outcome of getResponse()
        -userEntries is the list (array) that stores all of the user's entries
        -catNames is a list of the KeywordCategory names
        -catThresh is a list of the KeywordCategory thresholds
        -PATH is the filepath where the files that are read/written to are stored
    '''
    global counter, kw_dicts, STAGE, userEntries, catNames, catThresh, PATH, responseSpreadsheet
    backUp = []  # for resetting list when 'n' is entered
    # set up dictionaries by name
    for i in range(0, len(kw_dicts)):
        kw_dicts[i].set_name(catNames[i])
        kw_dicts[i].set_thresh(catThresh[i])

    # we are not in first run of function, and previous list has been confirmed
    if 0 < counter <= len(kw_dicts) and userEntries[-2].lower() == 'y':
        keywords = userEntries[-1]
        kw_dicts[(counter - 1)].set_list(keywords)
        kw_dicts[(counter - 1)].set_dict(kw_dicts[(counter - 1)].getKeyList())

    # User made an error with one of the keywords, retry
    if 0 < counter <= len(kw_dicts) + 1 and userEntries[-1].lower() == 'n':
        userEntries.pop()
        userEntries.pop()
        counter -= 1
        getResponse(STAGE)
    # If previous keywords have been confirmed, move on to next
    elif counter < len(kw_dicts) and userEntries[-1].lower() == 'y':
        convo.insert(tk.END, f'\nComp: Enter the keywords of the category named "{kw_dicts[counter].getName()}"'
                                f'\nComp: (In the form "keyword1,keyword2" or "key word1,key word2" for multi-word keywords):')
        counter += 1
    # If the user accepts the final keyword category, jump to next step
    elif counter == len(kw_dicts) and userEntries[-1].lower() == 'y':
        counter += 1
        getResponse(STAGE)
    # Check that keywords are right each time, = allows us to recheck on the last keyword category
    elif counter <= len(kw_dicts):
        convo.insert(tk.END, f'\nComp: The keywords of the category named "{kw_dicts[(counter-1)].getName()}"'
                                f' are {kw_dicts[(counter-1)].getKeyList()}.'
                                f'\nComp: Is this correct? (Enter "Y" to confirm or "N" to change)')
        # If counter of categories is on the final category, increment it one more time to escape infinite loop

    # At this point the dictionaries are complete. Ask user if they would like  to create a backup of these keywords.
    elif counter == (len(kw_dicts)+1):
        for i in range(0, len(kw_dicts)):
            convo.insert(tk.END, f'\nComp: The keywords of the category named "{kw_dicts[i].getName()}"'
                                    f' are {kw_dicts[i].getKeyList()}.')
        # This is to save keywords to file for later use.

        convo.insert(tk.END, f'\nComp: Would you like to save these keywords to a text file for later use? '
                                f'\nComp: (Enter "Y" to save or "N" to skip):')
        counter += 1
    elif counter > (len(kw_dicts)+1) and userEntries[-1].lower() == 'y':
        '''if user said yes, open and write keywords to "keywords.txt"
        Format of "keywords.txt"
        ***************************************
        Category name 1
        Category 1 threshold
        Keyword 1,Keyword 2,Keyword 3
        Category name 2
        Category 2 threshold
        Keyword 4,Keyword 5,Keyword 6
        **************************************
        Notice spacing between "," and "K" between each keyword!
            This allows for multi-word keywords
        '''
        filename = 'keywords.txt'
        fullname = PATH + '' + filename
        file = open(fullname, 'w')
        # write keywords to text file using KeywordCategory class's built in method: textFileOutput
        for i in range(0, len(kw_dicts)):
            file.write(kw_dicts[i].textFileOutput())
        convo.insert(tk.END, f'\nComp: The keywords have been saved to file "{filename}" in the folder "{PATH}".\n')
        # counter will be reused, reset it.
        counter = 0
        STAGE = 2
        getResponse(STAGE)

    # if user does not want a backup, just move to next step.
    elif counter > (len(kw_dicts) + 1) and userEntries[-1].lower() == 'n':
        convo.insert(tk.END, f'\nComp: The keywords will not be saved.')
        counter = 0
        STAGE = 2
        getResponse(STAGE)
    else:
        convo.insert(tk.END, f"\nComp: I'm sorry, that entry was invalid. Please try again.")
        for i in range(0, len(userEntries)-1):
            backUp.append(userEntries[i])
        userEntries.clear()
        userEntries = backUp.copy()
        counter -= 1
        getResponse(STAGE)


def getTextToAnalyze():
    '''
    This function is for getting the (student's) text to be analyzed for keywords.
    This is achieved through manual entry (readCSV = false) or by reading from a csv file.
    The csv file is assumed to be a quiz export from canvas named "quizExport.csv" unless the .xlsx is read.
    If the .xlsx is read, manual = False, and logic is a little bit different.
    It uses the global variables: textToBeAnalyzed, counter, userEntries, STAGE, kw_dicts, manual, canvas_PATH, df, and readCSV.
        -textToBeAnalyzed is a string that stores the student's text after it has been entered into
            the textfield by the user
        -counter is a counter variable to count how many times we have called this function
        -userEntries is the list (array) that stores the users text entries.
        -STAGE is to be updated on the last run of the function so that we can go to the next step
        -kw_dicts is the list containing the KeywordCategory objects
        -manual is the boolean that determines manual or text file input
        -canvas_PATH is the filepath where the canvas files that are read/written to are stored.
        -df is the dataframe created from the inputted .xlsx file
        -readCSV is a boolean that is used in logic for if the user wants to manually enter keywords but
          read in text to be analyzed.
    '''
    global textToBeAnalyzed, counter, userEntries, STAGE, kw_dicts, manual, canvas_PATH, df, readCSV
    backUp = []  # For backup & resetting the userEntries list.

    # If manual entry was used in Stage 1
    if manual:
        if counter == 0:
            convo.insert(tk.END,
                         f'\n\nComp: Would you like to read text to be analyzed from a csv file, or manually input it? '
                         f'\nComp:(Enter "0" for csv file or "1" for manual entry).')
            counter += 1

        # Logic for manual or read text entry
        elif counter == 1:
            if userEntries[-1] == '0':
                readCSV = True
                counter += 1
            elif userEntries[-1] == '1':
                readCSV = False
                counter += 1
            else:
                convo.insert(tk.END, f'\n\nComp: That entry was invalid. Please try again.')

        # Manual text entry
        if not readCSV:
            if counter == 2:
                convo.insert(tk.END, f'\n\nComp: Please type or paste the text into the textbox '
                                     f'and press Enter when finished.\n')
                counter += 1
            elif counter == 3:
                textToBeAnalyzed = userEntries[-1]
                convo.insert(tk.END, f'\n\nComp: You entered:\n"{textToBeAnalyzed}"'
                                     f'\n\nComp: Is this correct? (Enter "Y" to confirm or "N" to go back)')
                counter += 1
            # Move to text analysis Stage
            elif counter == 4 and userEntries[-1].lower() == 'y':
                counter = 0
                STAGE = 3
                getResponse(STAGE)
            # Text entered has an error, go back and try again.
            elif counter == 4 and userEntries[-1].lower() == 'n':
                for i in range(0, len(userEntries) - 2):
                    backUp.append(userEntries[i])
                userEntries.clear()
                userEntries = backUp
                counter = 2
                getResponse(STAGE)

        # If entry by csv file is desired
        if readCSV:
            # filename the computer is expecting for the csv file (since this is to be used with Canvas)
            filename = 'quizExport.csv'

            if counter == 2:
                convo.insert(tk.END,
                             f'\nComp: Is the text you would like to analyze in: "{filename}" in the directory "{canvas_PATH}"?'
                             f'\nComp: (Enter "Y" to confirm or "N" to reject)')
                counter += 1

            if counter == 3 and userEntries[-1].lower() == 'y':
                # The "unique string" is used to retrieve the specific column in the Canvas CSV export.
                convo.insert(tk.END, f'\nComp: Enter the unique string that is part of your Question:')
                counter += 1

            elif counter == 4 and userEntries[-1].lower() != 'y':
                '''
                  This part of the function searches the csv column headers for a header that
                    contains the unique string entered by the user. This distinguishes one question 
                    from another (in the event of multiple essay response questions in one quiz).
                '''
                uniqueString = userEntries[-1].lower()
                print(f'Unique String: {uniqueString}')
                saved_column = None
                try:
                    fullname = canvas_PATH + '' + filename
                    df2 = pd.read_csv(fullname)
                    for column in df2:
                        print(f'column: {column}')
                        if uniqueString.lower() in column.lower():
                            columnName = column
                            saved_column = df2[columnName]
                            break
                        else:
                            saved_column = ''

                    if len(saved_column) <= 0:
                        convo.insert(tk.END, f'\nComp: The unique String could not be found. Please try again.')
                        counter -= 1
                        revert(1)
                        getResponse(STAGE)

                    if len(saved_column) >= 1:
                        convo.insert(tk.END, f'\nComp: From File:')
                        try:
                            for i in range(0, len(saved_column)):
                                print(textToBeAnalyzed)
                                if i == len(saved_column) - 1:
                                    textToBeAnalyzed += str(saved_column[i]).lower()
                                else:
                                    textToBeAnalyzed += str(saved_column[i]).lower() + '*'
                                convo.insert(tk.END, f'\nComp: Student {i + 1}:'
                                                     f'\n{saved_column[i]}\n')

                        except TypeError or ValueError:
                            for i in range(0, len(userEntries) - 1):
                                backUp.append(userEntries[i])
                            userEntries.clear()
                            userEntries = backUp.copy()
                            counter -= 1
                            getResponse(STAGE)
                        convo.insert(tk.END, f'\nComp: Is this correct? (Enter "Y" to confirm or "N" to change)')
                        counter += 1

                except FileNotFoundError:
                    convo.insert(tk.END, f'\nComp: File "{filename}" not found in "{PATH}"')
                    for i in range(0, len(userEntries) - 1):
                        backUp.append(userEntries[i])
                    userEntries.clear()
                    userEntries = backUp.copy()
                    counter -= 1
                    getResponse(STAGE)

            # If the text is not in the right file, tell user to put it there.
            elif counter == 3 and userEntries[-1].lower() == 'n':
                fullname = PATH + '' + filename
                convo.insert(tk.END,
                             f'\nComp: Please put the text to analyze into "{fullname}" to read from a text file\n')
                revert(1)
                counter -= 1
                getResponse(STAGE)

            # Entries are correct, move to analysis stage
            elif counter == 5 and userEntries[-1].lower() == 'y':
                counter = 0
                STAGE = 3
                getResponse(STAGE)

            # If text is wrong in text file, tell them to update it then try again.
            elif counter == 5 and userEntries[-1].lower() == 'n':
                convo.insert(tk.END,
                             f'\nComp: Please correct the text to analyze in "{filename}" to read it from a text file\n')
                revert(1)
                textToBeAnalyzed = ''
                counter -= 2
                getResponse(STAGE)

    # This is the part for entries being read fro the .xlsx file
    if not manual:
        if counter == 0:
            filename = df['Response'][1]
            counter += 1
            getResponse(STAGE)

            if counter == 1:
                '''
                    Assumed textToBeAnalyzed file format:
                    ********************************
                    Text that someone has wrote where
                    we are looking for keywords...
                    *
                    Second piece of text that we are
                    analyzing for keywords
                    ********************************
                '''
                uniqueString = df['Response'][2]
                try:
                    fullname = canvas_PATH+''+filename
                    df2 = pd.read_csv(fullname)
                    for column in df2:
                        if uniqueString in column:
                            columnName = column
                            saved_column = df2[columnName]
                    #convo.insert(tk.END, f'\nComp: From File:')  # this can be used for debugging
                    try:
                        for i in range(0, len(saved_column)):
                            if i == len(saved_column)-1:
                                textToBeAnalyzed += str(saved_column[i]).lower()
                            else:
                                textToBeAnalyzed += str(saved_column[i]).lower() + '*'
                            #convo.insert(tk.END, f'\nComp: Student {i + 1}:'  # this can also be
                            #                     f'\n{saved_column[i]}\n')    # used for debugging
                        counter = 0
                        STAGE = 3
                        getResponse(STAGE)
                    except ValueError or TypeError:
                        for i in range(0, len(userEntries) - 1):
                            backUp.append(userEntries[i])
                        userEntries.clear()
                        userEntries = backUp.copy()
                        counter -= 1
                        getResponse(STAGE)
                    counter += 1
                # File is not in proper directory. We should never hit this because of preprocess(), but to be safe.
                except FileNotFoundError:
                    convo.insert(tk.END, f'\nComp: File "{filename}" not found in "{PATH}". '
                                         f'Please move the file to the directory or confirm you have entered the correct filename into the spreadsheet.\n')
                    for i in range(0, len(userEntries) - 1):
                        backUp.append(userEntries[i])
                    userEntries.clear()
                    userEntries = backUp.copy()
                    counter = 3
                    STAGE = 3
                    getResponse(STAGE)


def analyzeText():
    '''
    This function does the actual analysis of the inputted text and can also export results to a csv file
    It uses the global variables: userEntries, textToBeAnalyzed kw_dicts, counter, STAGE, catNames,
     catThresh, csvInput, PATH, manual, scores, assignmentName, and df.
        -userEntries is the list of user input
        -textToBeAnalyzed is the text the user inputted in the previous step for the program to analyze
        -kw_dicts is the list of KeywordCategory objects
        -counter is a counter variable to count how many times we have run this function
        -STAGE is used to track with "stage" of the program we are on and updated to move back or forward
        -csvInput is a list that stores the counts of each KeywordCategory to be output to a csv file
        -PATH is the filepath where the files that are read/written to are stored
        -manual is used in this function to reset it to None if the user has more text to analyze.
        -scores is the list that will hold the student's scores on their text entries.
        -assignmentName is the name of the assignment that the scores are to be recorded to in a Canvas gradebook export
        -df is the dataframe storing entries from the prepared .xlsx file
    '''
    global userEntries, textToBeAnalyzed, kw_dicts, counter, STAGE, csvInput, PATH, manual, scores, assignmentName, df
    backUp = []
    # If the user is using the .xlsx file, these entries will save the Yes/No of saving statistics,
    #   saving scores(grades), and the name of the assignment to save the grades to.
    if manual:
        if counter == 0:
            studentEntries = textToBeAnalyzed.split('*')
            print(studentEntries)
            for i in range(0, len(studentEntries)):
                convo.insert(tk.END, f'\n\n\nComp: Text Analysis for Student #{i + 1}:\n{LINE_SPACE}')
                kw_dictList, categoryCounts = findAllKeywords(studentEntries[i], kw_dicts)
                allMatches = 0
                allPassFails = []
                passedAll = True
                for i in range(0, len(kw_dictList)):
                    allMatches += categoryCounts[i]
                    allPassFails.append(kw_dicts[i].getPassFail())
                    convo.insert(tk.END, f'\nCategory:"{kw_dicts[i].getName()}",'
                                         f'\nThreshold:{kw_dicts[i].getThresh()},'
                                         f' Total keywords found: {categoryCounts[i]}'
                                         f'\nOccurrences of each Keyword: {kw_dictList[i]}\n'
                                         f'\nPassed Threshold: {kw_dicts[i].getPassFail()}\n')

                categoryCounts.append(allMatches)

                for i in range(0, len(allPassFails)):
                    passedAll = passedAll and allPassFails[i]
                    # If student passes all thresholds, they will be given a mark of 1.
                if passedAll:
                    scores.append(1)
                    categoryCounts.append(True)
                # Otherwise they will be given a mark of 0.
                else:
                    scores.append(0)
                    categoryCounts.append(False)
                csvInput.append(categoryCounts)
                convo.insert(tk.END, f'\nTotal Keyword Matches: {allMatches}'
                                     f'\nPassed All Thresholds: {passedAll}\n{LINE_SPACE}\n')
            counter += 1
            getResponse(STAGE)

        elif counter == 1:
            convo.insert(tk.END,
                         f'\nComp: Would you like to save these statistics to a CSV file? (Enter "Y" for Yes or "N" for No)')
            counter += 1

        elif counter == 2 and userEntries[-1].lower() == 'y':
            '''
                        if user said yes, open and write keywords to "textAnalysis.csv"
                            Format of "textAnalysis.csv"
                            **********************************************************
                            Matches Category 1, Matches Category 2, ..., Total Matches
                                Text 1 Matches, Text 1 Matches,     ..., Text 1 TotalMatches 
                                Text 2 Matches, Text 2 Matches,     ..., Text 2 TotalMatches
                            **********************************************************

                            This should give us an excel file like this:
                            **********************************************************************
                            | Matches Category 1 | | Matches Category 2 |    | Total Matches |
                            ---------------------- ---------------------- ------------------------
                            | Text 1 Matches C1  | | Text 1 Matches C2  | | Text 1 Total Matches |
                            ---------------------- ---------------------- ------------------------
                            **********************************************************************
                        '''
            # This code writes the number of matches for each KeywordCategory and the total number of
            #  matches to the "textAnalysis.csv" file
            csvHeaders = []
            for i in range(0, len(kw_dicts)):
                csvHeaders.append(f'Matches {kw_dicts[i].getName()} ({kw_dicts[i].getThresh()})')
            csvHeaders.append('Total Matches')
            csvHeaders.append('Passed (T/F)')
            filename = 'textAnalysis.csv'
            fullname = PATH + '' + filename
            file = open(fullname, 'w', newline='')
            # write keywords to text file using KeywordCategory class's built in method: textFileOutput
            writer = csv.writer(file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL, dialect='excel')
            writer.writerow(csvHeaders)
            for i in range(0, len(csvInput)):
                writer.writerow(csvInput[i])
            convo.insert(tk.END,
                         f'\n\nComp: The analytics have been saved to file "{filename}" in the folder "{PATH}".')
            convo.insert(tk.END,
                         f'\nComp: Would you like to save student scores (pass/fail) to a Canvas gradebook export? (Enter "Yes" or "No").')
            counter += 1
            getResponse(STAGE)

        elif counter == 2 and userEntries[-1].lower() == 'n':
                convo.insert(tk.END, f'\nComp: Statistics will not be saved.')
                convo.insert(tk.END, f'\nComp: Would you like to save student scores (pass/fail) to a Canvas gradebook export? (Enter "Y" for Yes or "N" for No).')
                counter += 1

        elif counter == 3 and userEntries[-1].lower() == 'y':
            # This is used to retrieve the column name in the Canvas CSV export.
            writeToCanvasCSV(assignmentName)
            counter += 1
            getResponse(STAGE)

        elif counter == 3 and userEntries[-1].lower() == 'n':
            convo.insert(tk.END, f'\nComp: Canvas statistics will not be saved.')
            counter += 1
            getResponse(STAGE)

        # Since they are manually entering one at a time, ask if user has more text to analyze
        elif counter == 4:
            convo.insert(tk.END, f'\nComp: Is there more text you would like to analyze? (Enter "C" to continue or "Exit" to Exit Program):')
            counter += 1

        # Resets ALL variables so the program can be re-run
        elif counter == 5 and userEntries[-1].lower() == 'c':
            convo.insert(tk.END, f'\nComp: Program Resetting...\n\n')
            reset()

        # If no more text to analyze, Output Goodbye Message, Exit program
        elif counter == 5 and userEntries[-1].lower() == 'exit':
            convo.insert(tk.END, f'\nComp: Thank you for using KeywordFinder, Goodbye!')
            print('Thank you for using KeywordFinder, Goodbye!')
            time.sleep(3)
            counter += 1
            getResponse(STAGE)

        # Exits the program
        elif counter == 6 and userEntries[-1].lower() == 'exit':
            raise SystemExit(0)

        else:
            convo.insert(tk.END, "\nComp: I'm sorry, that entry was invalid.")
            print(counter)
            print('ue: ', userEntries)
            for i in range(0, len(userEntries) - 1):
                backUp.append(userEntries[i])
            userEntries.clear()
            userEntries = backUp.copy()
            counter -= 1
            getResponse(STAGE)


    if not manual:
        saveStats = df['Response'][3]
        saveScores = df['Response'][4]
        assignmentName = df['Response'][5]
        if counter == 0:
            studentEntries = textToBeAnalyzed.split('*')
            print(studentEntries)
            for i in range(0, len(studentEntries)):
                convo.insert(tk.END, f'\n\n\nComp: Text Analysis for Student #{i + 1}:\n{LINE_SPACE}')
                kw_dictList, categoryCounts = findAllKeywords(studentEntries[i], kw_dicts)
                allMatches = 0
                allPassFails = []
                passedAll = True
                for i in range(0, len(kw_dictList)):
                    allMatches += categoryCounts[i]
                    allPassFails.append(kw_dicts[i].getPassFail())
                    convo.insert(tk.END, f'\nCategory:"{kw_dicts[i].getName()}",'
                                            f'\nThreshold:{kw_dicts[i].getThresh()},'
                                            f' Total keywords found: {categoryCounts[i]}'
                                            f'\nOccurrences of each Keyword: {kw_dictList[i]}\n'
                                            f'\nPassed Threshold: {kw_dicts[i].getPassFail()}\n')

                categoryCounts.append(allMatches)


                for i in range(0, len(allPassFails)):
                    passedAll = passedAll and allPassFails[i]
                    # If student passes all thresholds, they will be given a mark of 1.
                if passedAll:
                    scores.append(1)
                    categoryCounts.append(True)
                # Otherwise they will be given a mark of 0.
                else:
                    scores.append(0)
                    categoryCounts.append(False)
                csvInput.append(categoryCounts)
                convo.insert(tk.END, f'\nTotal Keyword Matches: {allMatches}'
                                        f'\nPassed All Thresholds: {passedAll}\n{LINE_SPACE}')
            counter += 1
            getResponse(STAGE)

        elif counter == 1 and saveStats.lower() == 'no':
            counter += 1
            getResponse(STAGE)

        # Export to csv file
        elif counter == 1 and saveStats.lower() == 'yes':
            '''
            if user said yes, open and write keywords to "textAnalysis.csv"
                Format of "textAnalysis.csv"
                **********************************************************
                Matches Category 1, Matches Category 2, ..., Total Matches
                    Text 1 Matches, Text 1 Matches,     ..., Text 1 TotalMatches 
                    Text 2 Matches, Text 2 Matches,     ..., Text 2 TotalMatches
                **********************************************************
                    
                This should give us an excel file like this:
                **********************************************************************
                | Matches Category 1 | | Matches Category 2 |    | Total Matches |
                ---------------------- ---------------------- ------------------------
                | Text 1 Matches C1  | | Text 1 Matches C2  | | Text 1 Total Matches |
                ---------------------- ---------------------- ------------------------
                **********************************************************************
            '''
            # This code writes the number of matches for each KeywordCategory and the total number of
            #  matches to the "textAnalysis.csv" file
            csvHeaders = []
            for i in range(0, len(kw_dicts)):
                csvHeaders.append(f'Matches {kw_dicts[i].getName()} ({kw_dicts[i].getThresh()})')
            csvHeaders.append('Total Matches')
            csvHeaders.append('Passed (T/F)')
            filename = 'textAnalysis.csv'
            fullname = PATH+''+filename
            file = open(fullname, 'w', newline='')
            # write keywords to text file using KeywordCategory class's built in method: textFileOutput
            writer = csv.writer(file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL, dialect='excel')
            writer.writerow(csvHeaders)
            for i in range(0, len(csvInput)):
                writer.writerow(csvInput[i])
            convo.insert(tk.END, f'\n\nComp: The analytics have been saved to file "{filename}" in the folder "{PATH}".')
            counter += 1
            getResponse(STAGE)

        elif counter == 2 and saveScores.lower() == 'no':
            convo.insert(tk.END, f'\nComp: Canvas statistics will not be saved.')
            counter += 1
            getResponse(STAGE)

        elif counter == 2 and saveScores.lower() == 'yes':
            # This is used to retrieve the column name in the Canvas CSV export.
            writeToCanvasCSV(assignmentName)
            counter += 1
            getResponse(STAGE)

        # Since they are manually entering one at a time, ask if user has more text to analyze
        elif counter == 3:
            convo.insert(tk.END, f'\nComp: Is there more text you would like to analyze? (Enter "C" to continue or "Exit" to Exit Program):')
            counter += 1

        # Resets ALL variables so the program can be re-run
        elif counter == 4 and userEntries[-1].lower() == 'c':
            convo.insert(tk.END, f'\nComp: Program Resetting...\n\n')
            reset()

        # If no more text to analyze, Output Goodbye Message, Exit program
        elif counter == 4 and userEntries[-1].lower() == 'exit':
            convo.insert(tk.END, f'\nComp: Thank you for using KeywordFinder, Goodbye!')
            print('Thank you for using KeywordFinder, Goodbye!')
            time.sleep(3)
            counter += 1
            getResponse(STAGE)

        # Exits the program
        elif counter == 5 and userEntries[-1].lower() == 'exit':
            raise SystemExit(0)

        else:
            convo.insert(tk.END, "\nComp: I'm sorry, that entry was invalid.")
            print(counter)
            print('ue: ', userEntries)
            for i in range(0, len(userEntries) - 1):
                backUp.append(userEntries[i])
            userEntries.clear()
            userEntries = backUp.copy()
            counter -= 1
            getResponse(STAGE)


def findAllKeywords(text, kw_dicts):
    '''
    This function runs each category dictionary in our kwDictList array
    Through the existing findKeywords method
    It updates the count of the dictionary at each index
    and returns a count of keywords from that category
    which gets stored in the categoryCounts array.
    '''

    kwDictList = ['']*len(kw_dicts)
    categoryCounts = [0]*len(kw_dicts)
    for i in range(0, len(kw_dicts)):
        kw_dicts[i].resetKeys()
        kwDictList[i] = kw_dicts[i].getKeywords()

    for i in range(0, len(kwDictList)):
        kwDictList[i], categoryCounts[i] = findKeywords(text, kwDictList[i])
    return kwDictList, categoryCounts


def findKeywords(text, keywordDictionary):
    '''
    re library contains the split method to split the text into regular English words.
    This means it removes commas, hyphens, apostrophes...-> ","   "-"   " ' " .. etc.
    This means that the keywords cannot contain these characters
    OR something else must be used to split the text word-by-word (more difficult)
    '''
    # totalMatches keeps track of number of matches to any keyword
    totalMatches = 0
    # use the dictionary called to the method but copied because we will be updating it
    dictionary = keywordDictionary
    # this line splits the input text into individual words and puts that into an array
    # we also use the lowercase version to prevent capitalized letters from being skipped
    array = re.split('\W+', text.lower())

    for i in range(0,len(array)):
        # pulls next word to compare
        nextWord = array[i]
        # if we are not on the last word of the input text we can compare two words at a time
        # THIS IS IMPORTANT. This is how to find multi-word keywords like "design thinking"
        if i < len(array)-1:
            # next two words is just current word concatenated with a space and the next word
            nextTwoWords = nextWord+' '+array[i+1]
            # if the next two words are in our dictionary, they are a keyword
            if nextTwoWords in dictionary:
                # we increment the number at that keyword in the dictionary (so we can track KW count
                # for each word)
                dictionary[nextTwoWords] += 1
                totalMatches += 1
        '''this part just compares word by word.
            for this reason, one word keywords that are also contained in two word keywords are an issue
            Ex. "design thinking" and "thinking" cannot both be keywords because "design thinking" will be counted as 2 keywords
            because the text is being scanned for both single and double word keywords
            '''
        if nextWord in dictionary:
            # Same as above, if dictionary contains the word, it is a keyword.
            dictionary[nextWord] += 1
            totalMatches += 1
    # this lets us return both the dictionary (which now has how many times each keyword occurs)
    # and the total number of keywords.

    return dictionary, totalMatches


def writeToCanvasCSV(assignmentName):
    '''
    This function takes a CSV file exported from Canvas LMS, searches for an assignment name
      in it, and writes the scores from the text analysis under that column, then exports that
      to a new CSV file that can be imported back to Canvas to update the grades.
    This is done by traversing the export csv file to fin which column the assnment name is in,
      then converting that file to an xlsx (Excel spreadsheet) and writing the scores to the
      rows underneath that column. Then that xlsx spreadsheet is converted back into a CSV file
      and the xlsx file and temporary directory are deleted to prevent confusion of the user.
    '''
    global scores, counter, canvas_PATH, STAGE
    print(f'Write to CSV:{scores} ')
    csvName = 'canvasExport.csv'
    newFileName = 'canvasImport.csv'

    xlsxName = 'tempSpreadsheet.xlsx'
    tempPATH = canvas_PATH + 'temp/'
    newCSVFull = canvas_PATH + '' + newFileName

    csvFull = canvas_PATH + '' + csvName
    xlsxFull = tempPATH + '' + xlsxName

    # Note that "assignName" will come from user input and scores are the marks are based on
    #  the pass/fail assessment in the previous step.

    # This code finds which column we are looking for in the csv file
    try:
        file = open(csvFull, 'rt')
        reader = csv.reader(file, delimiter=',')
    except FileNotFoundError:
        convo.insert(tk.END, f'\nComp: The canvas export file "{csvName}" cold not be found in "{csvFull}". '
                             f'Please export the file from Canvas and move the file to the directory or confirm you have entered the correct filename into the spreadsheet.\n')

    else:
        colLet = 'A'  # first column in Excel
        colCount = ord(colLet)  # Get the ascii of our column in Excel
        for row in reader:
            for field in row:
                if assignmentName in field:
                    colLet = chr(colCount)  # converts ascii to character
                colCount += 1

        if not os.path.exists(tempPATH):
            os.makedirs(tempPATH)

        # This code writes the csv file above to an xslx file
        # Code modified from: https://stackoverflow.com/questions/17684610/python-convert-csv-to-xlsx
        wb = Workbook()
        ws = wb.active
        with open(csvFull, 'r') as f:
            for row in csv.reader(f):
                ws.append(row)
        wb.save(xlsxFull)

        # This code reads the xlsx file created above and writes the new grades to it.
        # Code modified from: https://stackoverflow.com/questions/49681392/python-pandas-how-to-write-in-a-specific-column-in-an-excel-sheet
        df = pd.DataFrame({assignmentName: scores})
        wb = load_workbook(xlsxFull)
        ws = wb['Sheet']
        # gap is the row number to start at (excel indexes from 1)
        # row 1 contains assignment name, row 2 contains max grade on assigment (Canvas default formatting)
        gap = 3
        for index, row in df.iterrows():
            cell = f'{colLet}%d' % (index + gap)
            ws[cell] = row[0]

        wb.save(xlsxFull)

        # This code writes the previously made and modified xlsx file to a CSV
        #  so that it can be imported into canvas.
        # Code modified from: https://stackoverflow.com/questions/20105118/convert-xlsx-to-csv-correctly-using-python
        wb = xlrd.open_workbook(xlsxFull)
        sh = wb.sheet_by_name('Sheet')
        csvFile = open(newCSVFull, 'w')
        wr = csv.writer(csvFile, delimiter=',', lineterminator='\n')

        for rownum in range(sh.nrows):
            wr.writerow(sh.row_values(rownum))

        csvFile.close()

        # This code removes the temporary files we created.
        os.remove(xlsxFull)  # we must remove files from directory before the directory itself
        os.rmdir(tempPATH)  # delete folder "temp"
        # os.remove(csvFull)  # optional: delete the original Canvas Export file so that there is only 1 file at the end.

        convo.insert(tk.END, f'\nComp: The student scores have been saved to file "{newFileName}" in the folder "{newCSVFull}".\n')

def revert(n):
    global userEntries, STAGE
    backUp = []
    for i in range(0, len(userEntries) - n):
        backUp.append(userEntries[i])
    userEntries.clear()
    userEntries = backUp.copy()
    getResponse(STAGE)


def reset():
    global userEntries, counter, numCat, catNames, catThresh, kw_dicts, STAGE, textToBeAnalyzed, manual, csvInput, assignmentName, scores, responseSpreadsheet, df

    userEntries = []
    counter = 0
    numCat = 0
    catNames.clear()
    catThresh.clear()
    kw_dicts.clear()
    STAGE = 0
    textToBeAnalyzed = ''
    manual = None
    csvInput.clear()
    assignmentName = ''
    scores.clear()
    responseSpreadsheet = ''
    df = None
    #getResponse(STAGE)


def getColour(field, colourmode):
    '''
    FFFFFF is the hex colour for white, 000000 is the hex color for black
    darkmode is white text on black background
    lightmode is black text on white background
    field is what they are colouring, bg = background colour
    txt is all text in the program
    '''
    if colourmode == 'dark':
        if field == 'bg':
            return '#000000'
        elif field == 'txt':
            return '#FFFFFF'
        elif field == 'scrl':
            return '#000000'
        elif field == 'frm':
            return '#A7A7A7'
    if colourmode == 'light':
        if field == 'bg':
            return '#FFFFFF'
        elif field == 'txt':
            return '#000000'
        elif field == 'scrl':
            return '#FFFFFF'
        elif field == 'frm':
            return '#000000'


def setColour(colourmode):
    '''
    FFFFFF is the hex colour for white, 000000 is the hex color for black
    darkmode is white text on black background
    lightmode is black text on white background
    field is what they are colouring, bg = background colour
    txt is all text in the program
    '''
    global bgcolour, textcolour, scrlcolour, framecolour
    bgcolour = getColour('bg', colourmode)
    textcolour = getColour('txt', colourmode)
    scrlcolour = getColour('scrl', colourmode)
    framecolour = getColour('frm', colourmode)
    canvas.config(bg=bgcolour)
    header_frame.config(bg=bgcolour)
    title.config(bg=bgcolour, fg=textcolour)
    credit.config(bg=bgcolour, fg=textcolour)
    middle_frame.config(bg=framecolour)
    lower_frame.config(bg=bgcolour)
    convo.config(bg=scrlcolour, fg=textcolour)
    user.config(bg=bgcolour, fg=textcolour)

def clear():
    convo.config(state='normal')
    convo.delete('12.0', tk.END)
    reset()
    getResponse(STAGE)



window = tk.Tk()

colourmode = 'light'
bgcolour = getColour('bg', colourmode)
textcolour = getColour('txt', colourmode)
scrlcolour = getColour('scrl', colourmode)
framecolour = getColour('frm', colourmode)

window.iconbitmap(default=iconpath)
window.title(f'KeywordFinder v{version}')

canvas = tk.Canvas(window, height=HEIGHT, width=WIDTH, bg=bgcolour)
canvas.pack(fill='both', expand=True)

header_frame = tk.Frame(window, bg=bgcolour, bd=6)
header_frame.place(anchor='n', relx=0.5, rely=0.001, relwidth=0.98, relheight=0.25)

title = tk.Label(header_frame, text='Keyword Finder', font=('Times', 40), bg=bgcolour, fg=textcolour)
title.place(anchor='n', relx=0.5, rely=0, relwidth=1, relheight=0.35)

credit = tk.Label(header_frame, text='Created by Daulton Baird', font=('Times', 12), bg=bgcolour, fg=textcolour)
credit.place(anchor='n', relx=0.5, rely=0.36, relwidth=1, relheight=0.12)

lmButton = ttk.Button(header_frame, text='LightMode', command=lambda: setColour('light'))
lmButton.place(relx=0.04, rely=0.02, relwidth=0.07, relheight=0.2)

dmButton = ttk.Button(header_frame, text='DarkMode', command=lambda: setColour('dark'))
dmButton.place(relx=0.12, rely=0.02, relwidth=0.07, relheight=0.2)

clrButton = ttk.Button(header_frame, text='Reset', command=lambda: clear())
clrButton.place(relx=0.90, rely=0.02, relwidth=0.07, relheight=0.2)

middle_frame = tk.Frame(window, bg=framecolour, bd=6)
middle_frame.place(relx=0.01, rely=0.17, relwidth=0.98, relheight=0.6)

# code from: https://stackoverflow.com/questions/17657212/how-to-code-the-tkinter-scrolledtext-module
scrollbar = tk.Scrollbar(middle_frame)
convo = tk.Text(middle_frame, width=10, height=10, wrap="word", yscrollcommand=scrollbar.set,
                borderwidth=0, bg=bgcolour, fg=textcolour, font=16, )
convo.insert(tk.END, welcome_message)
convo.see(tk.END)
getResponse(STAGE)
scrollbar.config(command=convo.yview)
scrollbar.pack(side="right", fill="y")
convo.pack(side="left", fill="both", expand=True)
convo.config(state='disable')

lower_frame = tk.Frame(window, bg=bgcolour, bd=6)
lower_frame.place(anchor='sw', relx=0.01, rely=0.99, relwidth=0.98, relheight=0.11)

user = tk.Entry(lower_frame, bg=bgcolour, font=40, fg=textcolour)
user.place(relx=0, rely=0.2, relwidth=0.7, relheight=0.7)
user.bind('<Return>', lambda event: updateText(user.get()))

button = ttk.Button(lower_frame, text='Enter', command=lambda: updateText(user.get()))
button.place(relx=0.75, rely=0.2, relwidth=0.25, relheight=0.7)

window.mainloop()
